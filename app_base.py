# ui/app.py
import os
import time
import sys
import tempfile
import shutil
import requests
print(requests.get("https://api.ipify.org").text)

from collections import deque
from pathlib import Path
from urllib.parse import urljoin
from datetime import datetime
from openpyxl import Workbook, load_workbook

from PyQt5.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QPushButton, QLineEdit, QListWidget,
    QLabel, QMessageBox, QApplication, QProgressDialog, QTextEdit, QMessageBox, QComboBox
)
from PyQt5.QtCore import QTimer, QDateTime

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, WebDriverException,
    ElementClickInterceptedException, UnexpectedAlertPresentException,
    NoAlertPresentException
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# suas funções internas

from ui.alert_menssenger import mostrar_toast
from .vpn_manager import VPNManager


def pasta_app():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent.parent

BASE_DIR = pasta_app()

def resource_path(relative):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath("."), relative)

PROFILE_PATH = BASE_DIR / "ui"/ "chrome_profile"

EXCEL_PROCESSOS_PATH = BASE_DIR / "processos.xlsx"


DOWNLOAD_DIR = BASE_DIR / "pdfs"
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

from openpyxl import Workbook

if not EXCEL_PROCESSOS_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Numero_Processo"
    wb.save(EXCEL_PROCESSOS_PATH)

# Chromedriver / Chrome options
HEADLESS = False  # não usar headless para extensões
CHROME_BINARY = None  # se usa caminho custom do navegador (opcional)

# Timeouts
WAIT_SHORT = 2
WAIT_MEDIUM = 8
WAIT_LONG = 25

# INPI URLs
URL_INPI = "https://busca.inpi.gov.br/pePI/"
URL_DESTINO = "https://busca.inpi.gov.br/pePI/jsp/marcas/Pesquisa_num_processo.jsp"

# -------------------------------------------

class SeleniumController:
    def __init__(self):
        self.driver = None

    def start(self):
        from selenium.webdriver.chrome.service import Service

        chrome_options = Options()
        profile_temp = tempfile.mkdtemp()
     
        
        chrome_options.add_argument(f"--user-data-dir={PROFILE_PATH}")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--autoplay-policy=no-user-gesture-required")
        prefs = {
            "download.default_directory": os.path.abspath(DOWNLOAD_DIR),
            "download.prompt_for_download": False,
            "plugins.always_open_pdf_externally": True,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        # Adicionar logs de erro verbose para debug
        chrome_options.add_argument("--verbose")
        chrome_options.add_argument("--no-first-run")
        # Definir flags de estabilização
        chrome_options.add_argument("--disable-software-rasterizer")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")

        # Criar novo serviço para o Selenium
        service = Service()
        service.start_timeout = 60

        # Aguardar estabilização de rede antes de iniciar o Chrome
        self._aguardar_rede_estavel()

        try:
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.set_window_size(600, 720)
            return self.driver
        except Exception as e:
            raise RuntimeError(f"Erro ao iniciar ChromeDriver: {e}")

    
    def _aguardar_rede_estavel(self, timeout=30):
        print("⏳ Aguardando rede estabilizar após a troca de VPN...")
        inicio = time.time()

        while time.time() - inicio < timeout:
            try:
                r = requests.get("https://www.google.com", timeout=5)
                if r.status_code == 200:
                    print("✅ Rede estabilizada.")
                    return
            except requests.exceptions.RequestException:
                pass
            time.sleep(3)  # Aguardar mais tempo
        raise RuntimeError("❌ A rede não estabilizou após a troca de VPN.")

    def stop(self):
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass

    def wait_for_element(self, by, value, timeout=WAIT_MEDIUM):
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))

    def wait_for_clickable(self, by, value, timeout=WAIT_MEDIUM):
        return WebDriverWait(self.driver, timeout).until(EC.element_to_be_clickable((by, value)))

    def safe_click(self, by, value, timeout=WAIT_MEDIUM):
        try:
            el = self.wait_for_clickable(by, value, timeout=timeout)
            el.click()
            return True
        except Exception as e:
            print(f"[selenium] safe_click falhou ({by},{value}): {e}")
            return False

    def element_exists(self, by, value):
        try:
            self.driver.find_element(by, value)
            return True
        except NoSuchElementException:
            return False

    def get_iframe_by_src_contains(self, partial_src):
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        for fr in iframes:
            try:
                src = fr.get_attribute("src") or ""
                if partial_src in src:
                    return fr
            except Exception:
                continue
        return None
    
    def popup_erro_download_visivel(self):
        try:
            popup = self.driver.find_element(
                By.XPATH,
                "//div[contains(text(),'Erro') or contains(text(),'erro')]"
            )
            return popup.is_displayed()
        except:
            return False
    def wait_for_download(self, timeout=90):
        """
        Aguarda o ÚLTIMO PDF realmente baixado (por data de modificação)
        """
        t0 = time.time()

        # snapshot inicial
        arquivos_iniciais = {
            f: os.path.getmtime(os.path.join(DOWNLOAD_DIR, f))
            for f in os.listdir(DOWNLOAD_DIR)
            if f.lower().endswith(".pdf")
        }

        while time.time() - t0 < timeout:
            if self.popup_erro_download_visivel():
                print("❌ Erro exibido pelo site durante download")
                raise RuntimeError("❌ Erro exibido pelo site durante download")
            pdfs = []

            for nome in os.listdir(DOWNLOAD_DIR):
                if not nome.lower().endswith(".pdf"):
                    continue

                caminho = os.path.join(DOWNLOAD_DIR, nome)

                try:
                    mtime = os.path.getmtime(caminho)
                except FileNotFoundError:
                    continue

                # só PDFs criados/modificados depois do clique
                if nome not in arquivos_iniciais or mtime > arquivos_iniciais.get(nome, 0):
                    if self._arquivo_estavel(caminho):
                        pdfs.append((mtime, caminho))

            if pdfs:
                # retorna o MAIS RECENTE
                pdfs.sort(key=lambda x: x[0], reverse=True)
                return pdfs[0][1]

            time.sleep(0.5)

        return None
    def _arquivo_estavel(self, caminho, tentativas=5, intervalo=0.5):
        """
        Verifica se o tamanho do arquivo não muda (download finalizado)
        """
        try:
            tamanho_anterior = -1
            for _ in range(tentativas):
                tamanho = os.path.getsize(caminho)
                if tamanho == tamanho_anterior:
                    return True
                tamanho_anterior = tamanho
                time.sleep(intervalo)
        except FileNotFoundError:
            return False
        return False   
       
class MainApp(QWidget):
    def __init__(self):
        super().__init__()
        self.tentativas_por_processo = {}
        self.MAX_TENTATIVAS = 2

        # 1️⃣ Flags e dados
        self.parar_loop = False
        self.loop_ativo = False
        self.processando = False

        self.indice_processo = 0
        self.total_processos = 0
        self.processos_extraidos = 0
        
        self.arquivo_concluidos = "processos_concluidos.txt"
        self.processos = []
        self.processos_concluidos = set()
        
        self.vpn = VPNManager()
      
       
        
        # 2️⃣ Criar UI PRIMEIRO
        self._criar_interface()
        self._iniciar_monitor_ip()
        # 3️⃣ Carregar dados DEPOIS
        self._carregar_processos_concluidos()
        self.atualizar_lista_processos()

        # 4️⃣ Filtrar concluídos
        self.processos = [
            p for p in self.processos
            if p not in self.processos_concluidos
        ]
           
    def _criar_interface(self):  
        self.setWindowTitle("INPI - Automação")
        self.setGeometry(120, 80, 300, 100)

        # UI simples (mantendo layout similar ao anterior)
        self.layout = QHBoxLayout(self)

        left = QVBoxLayout()
       # right = QVBoxLayout()
          # 🔐 VPN UI (AGORA VISÍVEL)
        self.label_ip = QLabel("🌍 IP: ---")
        self.label_ip.setStyleSheet("font-weight: bold; color: green;")
        self.btn_trocar = QPushButton("🔁 Trocar ip")
        self.btn_trocar.clicked.connect(self.trocar_vpn)

        left.addWidget(self.label_ip)
        left.addWidget(self.btn_trocar)

        # Conecta VPN ao iniciar
        self.atualizar_ip()
        self.combo_usuario = QComboBox()
        self.combo_usuario.addItem("Selecione o usuário INPI")
        self.combo_usuario.model().item(0).setEnabled(False)  # desabilita seleção
        self.combo_usuario.setCurrentIndex(0)
        
        self.input_usuario = QLineEdit()
        self.input_usuario.setPlaceholderText("Usuário INPI")
        self.input_usuario.setReadOnly(True)
        
        self.input_senha = QLineEdit()
        self.input_senha.setPlaceholderText("Senha INPI")
        self.input_senha.setEchoMode(QLineEdit.Password)
        self.input_senha.setReadOnly(True)
        
        left.addWidget(self.combo_usuario)
        left.addWidget(self.input_usuario)
        left.addWidget(self.input_senha)
        self.btn_iniciar = QPushButton("abrir site inpi")
        self.btn_iniciar.clicked.connect(self.iniciar_selenium)
        left.addWidget(self.btn_iniciar)

        self.lista_processos = QListWidget()
        self.lista_processos.itemClicked.connect(self.selecionar_processo)
        left.addWidget(QLabel("Processos"))
        left.addWidget(self.lista_processos)

        self.entry_processo = QLineEdit(); self.entry_processo.setPlaceholderText("N°Processo")
        left.addWidget(self.entry_processo)
        self.btn_buscar = QPushButton("iniciar extração")
        self.btn_buscar.clicked.connect(self.iniciar_processamento_em_lote)
        self.btn_parar = QPushButton("Parar")
        self.btn_parar.clicked.connect(self.parar_processamento)
        left.addWidget(self.btn_parar)

        #self.btn_buscar.clicked.connect(self.abrir_detalhe_processo)
        left.addWidget(self.btn_buscar)

        self.label_status = QLabel("Status: idle")
        left.addWidget(self.label_status)
        
        self.label_contador = QLabel("Processos extraídos: 0 / 0")
        self.label_contador.setStyleSheet("font-size: 12pt; font-weight: bold;")
        left.addWidget(self.label_contador)
        self.console_log = QTextEdit()
        self.console_log.setReadOnly(True)
        self.console_log.setStyleSheet("""
            background-color: white;
            color: black;
            font-family: Consolas;
            font-size: 11pt;
        """)
        left.addWidget(self.console_log)

        self.layout.addLayout(left, 2)
        
        self.selenium = SeleniumController()
        self.driver = None

        # data holders
        self.processo_atual_id = None
        
 
        self.lista_processos.setEnabled(True)
        self.carregar_usuarios_excel()

        for u in self.usuarios:
            self.combo_usuario.addItem(u["usuario"])
        
        self.combo_usuario.currentIndexChanged.connect(self.on_usuario_selecionado)
    
    def _iniciar_monitor_ip(self):
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.atualizar_ip)
        self.timer.start(5000)
        self.atualizar_ip()
    
    def atualizar_ip(self):
        ip = self.vpn.ip_atual()
        self.label_ip.setText(f"🌍 IP: {ip}")

    def trocar_vpn(self):
        print("BOTÃO CLICADO")
        self.label_ip.setText("🔄 Trocando VPN...")
        ip = self.vpn.conectar()
        self.label_ip.setText(f"🌍 IP: {ip}")

    def on_usuario_selecionado(self, index):
        if index == 0:
            self.input_usuario.clear()
            self.input_senha.clear()
            return
    
        user = self.usuarios[index - 1]  # 👈 deslocamento
        self.input_usuario.setText(user["usuario"])
        self.input_senha.setText(user["senha"])


    def parar_processamento(self):
        if not self.loop_ativo:
            return

        self.loop_ativo = False
        self.processando = False

        self.log("🛑 Processamento interrompido pelo usuário.")
        self.label_status.setText("Status: parado")
        self.lista_processos.setEnabled(True)
   
    def _carregar_processos_concluidos(self):
        if not os.path.exists(self.arquivo_concluidos):
            return

        with open(self.arquivo_concluidos, "r", encoding="utf-8") as f:
            for linha in f:
                numero = linha.strip()
                if numero:
                    self.processos_concluidos.add(numero)
    
            # carrega lista de processos
            #QTimer.singleShot(200, self.atualizar_lista_processos)
    def iniciar_processamento_em_lote(self):

        # 🔐 Verifica se o Chrome/Selenium está ativo
        if not self.driver or not self.selenium.driver:
            self.ui_warning(
                "Chrome não iniciado",
                "Abra o site do INPI antes de iniciar a extração."
            )
            self.log("⚠️ Tentativa de iniciar lote sem Chrome aberto.")
            return

        if not self.processos:
            self.ui_warning("Aviso", "Nenhum processo carregado.")
            return
        
        self.loop_ativo = True 
        
        self.lista_processos.setEnabled(False)
        
         # 🔥 CONVERTE LISTA EM FILA
        self.processos = deque(self.processos)
        
        self.total_processos = len(self.processos)
        self.processos_extraidos = 0
        self._atualizar_contador_ui()

        self.log("🚀 Iniciando processamento em lote...")
        

        self._processar_proximo()
    def carregar_usuarios_excel(self):
        self.usuarios = []  # lista de dicts

        caminho = BASE_DIR / "credenciais.xlsx"
        if not caminho.exists():
            self.log("⚠️ Arquivo credenciais.xlsx não encontrado.")
            return

        wb = load_workbook(caminho, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue

            self.usuarios.append({
                "usuario": str(row[0]).strip(),
                "senha": str(row[1]).strip()
            })
    def _processar_proximo(self):
        if self.processando:
            return

        if not self.processos:
            self.log("🏁 Fila finalizada")
            return

        self.processando = True

        # 🔥 REMOVE AQUI
        self.numero_atual = self.processos.popleft()
        self.tentativas_por_processo.setdefault(self.numero_atual, 0)
        self.log(f"📦 Processando processo {self.numero_atual}"
                 f"(tentativa {self.tentativas_por_processo[self.numero_atual] + 1})")

        try:
            self.abrir_detalhe_processo()
        except Exception as e:
            self.log(f"❌ Erro inesperado: {e}")
            self.processando = False
            QTimer.singleShot(0, self._processar_proximo)

    def selecionar_processo(self, item):
        if self.processando:
            return  # ignora clique durante processamento em lote

        numero = item.text()
        self.entry_processo.setText(numero)
        self.log(f"📌 Processo selecionado: {numero}")
   
    def _registrar_processo_concluido(self, numero):
        if not numero:
            self.log("⚠️ Tentativa de registrar processo concluído com número inválido.")
            return

        with open(self.arquivo_concluidos, "a", encoding="utf-8") as f:
            f.write(str(numero) + "\n")

        self.processos_concluidos.add(numero)

  
    def ui_warning(self, titulo, msg):
        QTimer.singleShot(
            0,
            lambda: QMessageBox.warning(self, titulo, msg)
        )

    def ui_error(self, titulo, msg):
        QTimer.singleShot(
            0,
            lambda: QMessageBox.critical(self, titulo, msg)
        )
    
    def log(self, mensagem):

        timestamp = QDateTime.currentDateTime().toString("HH:mm:ss")

        # console estilo VSCode
        self.console_log.append(f"[{timestamp}] {mensagem}")
        self.console_log.ensureCursorVisible()

        # status curto
        self.label_status.setText(f"Status: {mensagem}")


    def atualizar_lista_processos(self):
        self.lista_processos.clear()
        self.processos = []

        try:
            if not EXCEL_PROCESSOS_PATH.exists():
                raise FileNotFoundError(
                    f"Arquivo não encontrado: {EXCEL_PROCESSOS_PATH}"
                )

            wb = load_workbook(EXCEL_PROCESSOS_PATH, data_only=True)
            ws = wb.active

            header = ws["A1"].value
            if header != "Numero_Processo":
                raise ValueError(
                    "A coluna A deve se chamar 'Numero_Processo'"
                )

            vistos = set()

            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if not row or not row[0]:
                    continue

                numero = str(row[0]).strip()

                # ignora duplicados e concluídos
                if numero in vistos or numero in self.processos_concluidos:
                    continue
    
                vistos.add(numero)
                self.processos.append(numero)
                self.lista_processos.addItem(numero)

            self.total_processos = len(self.processos)
            self._atualizar_contador_ui()

            self.log(f"📄 {len(self.processos)} processos carregados do Excel.")
    
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erro Excel",
                f"Falha ao carregar processos do Excel"
            )
            self.log(f"Falha ao carregar processos do Excel")

    def iniciar_selenium(self):
        self.log("Iniciando Chrome + extensão...")
    
        try:
            if self.driver:
                self.selenium.stop()
                self.driver = None
    
            time.sleep(3)  # 🔥 importante após VPN
    
            self.driver = self.selenium.start()
            self.log("Chrome iniciado")
            self.driver.get(URL_INPI)
            #self.driver.get("https://api.ipify.org")
            print("IP do Chrome:", self.driver.find_element(By.TAG_NAME, "body").text)

        except Exception as e:
            self.btn_iniciar.setEnabled(True)
            print( self,
                "Erro",
                f"Falha ao iniciar Selenium:\n{e}")
            QMessageBox.critical(
                self,
                "Erro",
                f"Falha ao iniciar Selenium:\n{e}"
            )

    def login_inpi(self): 
        self.driver.get("https://busca.inpi.gov.br/pePI/") 
        time.sleep(3) 
        self.driver.find_element(By.NAME, "T_Login").send_keys( self.input_usuario.text() ) 
        self.driver.find_element(By.NAME, "T_Senha").send_keys( self.input_senha.text() ) 
        self.driver.find_element( By.XPATH, "//input[@type='submit' and contains(@value,'Continuar')]" ).click()
    def garantir_login(self):
        """
         Garante que o usuário esteja:
        1) Logado
        2) Na página de pesquisa por número de processo (NumPedido visível)
        """

        try:
             # 🔎 Já está na página correta?
            self.driver.find_element(By.NAME, "NumPedido")
            self.log("✅ Usuário já está logado e na página correta.")
            return
        except:
            pass

        self.log("🔐 Usuário não está na página correta. Garantindo login...")
       
        # ======================
        # LOGIN
        # ======================
        try:
            self.login_inpi()
            self.tratar_popups_login()
        except Exception as e:
            raise Exception(f"Falha no login: {e}")

        # ======================
        # AGUARDA PÁGINA PRINCIPAL
         # ======================
        WebDriverWait(self.driver, WAIT_MEDIUM).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        self.log("🔑 Login realizado. Acessando menu Marcas...")

           # ======================
        # CLICA NO MENU MARCAS
        # ======================
        try:
            menu_marcas = WebDriverWait(self.driver, WAIT_MEDIUM).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "area[href*='Pesquisa_num_processo.jsp']")
                 )
            )
       
            # ⚠️ area precisa ser clicada via JS
            self.driver.execute_script("arguments[0].click();", menu_marcas)
       
        except TimeoutException:
            raise Exception("❌ Não foi possível localizar o menu Marcas")

        # ======================
        # AGUARDA CAMPO NumPedido
        # ======================
        try:
            WebDriverWait(self.driver, WAIT_MEDIUM).until(
                EC.presence_of_element_located((By.NAME, "NumPedido"))
            )
            self.log("✅ Página de pesquisa por processo carregada com sucesso.")
        except TimeoutException:
            raise Exception("❌ Campo NumPedido não apareceu após acessar Marcas")
    
    def extrair_por_rotulo(self, rotulo):
        try:
            el = WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    f"//td[contains(normalize-space(),'{rotulo}')]/following-sibling::td[1]"
                ))
            )
            return el.text.strip()
        except TimeoutException:
            self.log(f"⏱ {rotulo} não encontrado")
            return ""

    def extrair_titular(self):
         return self.extrair_por_rotulo("Titular(1):")
    def extrair_dados_pagina(self):
      
        self.log("🧩 Iniciando extração estruturada dos dados do INPI")
        #self.log_pagina_atual("antes da extração")
        self.titular = self.extrair_titular()
       

        self.log("📋 Resumo da extração:")
        self.log(f"   Titular       : {self.titular or 'NÃO ENCONTRADO'}")
        
        self.log("✅ Extração finalizada com sucesso.")

    def tratar_popups_login(self, timeout=5):
        """
        Trata qualquer popup que apareça durante login:
        - alert JS
        - modal HTML
        - janela extra
        """
        # 🔔 ALERT JavaScript
        try:
            WebDriverWait(self.driver, timeout).until(EC.alert_is_present())
            alert = self.driver.switch_to.alert
            texto = alert.text
            alert.accept()
            self.log(f"⚠️ Alert fechado automaticamente: {texto}")
            time.sleep(1)
        except TimeoutException:
            pass
        except UnexpectedAlertPresentException:
            try:
                self.driver.switch_to.alert.accept()
                self.log("⚠️ Alert inesperado fechado")
            except Exception:
                pass
    
        # 🪟 JANELA EXTRA
        try:
            handles = self.driver.window_handles
            if len(handles) > 1:
                principal = handles[0]
                for h in handles[1:]:
                    self.driver.switch_to.window(h)
                    self.driver.close()
                    self.log("🪟 Popup de janela fechado")
                self.driver.switch_to.window(principal)
        except Exception:
            pass

        # 🧱 MODAL HTML (se existir)
        try:
            modal = self.driver.find_elements(
                By.XPATH,
                "//button[contains(.,'Fechar') or contains(.,'OK') or contains(.,'Continuar')]"
            )
            for btn in modal:
                if btn.is_displayed():
                    btn.click()
                    self.log("🧱 Modal HTML fechado")
                    time.sleep(1)
        except Exception:
            pass   
        
    def ui_toast(self, mensagem, tempo=2000):
        QTimer.singleShot(
            0,
            lambda m=mensagem, t=tempo: mostrar_toast(m, t)
        )
    
    def garantir_acesso_peticiones(self):
        try:
            self.log("🔍 Verificando necessidade de amplo acesso às petições...")
    
            link_amplo = WebDriverWait(self.driver, 1).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//a[contains(normalize-space(),'Clique aqui para ter acesso as petições')]"
                ))
            )

            self.log("🔐 Acesso restrito detectado. Solicitando liberação...")
    
            self.driver.execute_script("arguments[0].scrollIntoView(true);", link_amplo)
            time.sleep(0.5)
    
            link_amplo.click()

            # aguarda o modal abrir (título da página do popup)
            WebDriverWait(self.driver, 1).until(
                EC.title_contains("Finalidade do Acesso")
            )

            self.log("🪟 Modal de acesso aberto")
    
            # dependendo do INPI, basta abrir o modal para liberar
            # aguarda voltar para a página principal
            WebDriverWait(self.driver, 1).until(
                EC.not_(EC.title_contains("Finalidade do Acesso"))
            )

            self.log("✅ Acesso às petições liberado")

        except TimeoutException:
            # não existe bloqueio → segue normal
            self.log("🔓 Nenhum bloqueio de petições detectado")
    
    def liberar_acesso_peticiones(self):
        janela_principal = self.driver.current_window_handle

        try:
            self.log("🔍 Verificando popup de amplo acesso...")

            # aguarda abrir nova janela (popup)
            WebDriverWait(self.driver, 1).until(
                lambda d: len(d.window_handles) > 1
            )

            # muda para o popup
            for janela in self.driver.window_handles:
                if janela != janela_principal:
                    self.driver.switch_to.window(janela)
                    break

            self.log("🪟 Popup de amplo acesso detectado")

            # aguarda checkbox aparecer
            checkbox = WebDriverWait(self.driver, 1).until(
                EC.element_to_be_clickable((By.ID, "aceite"))
            )
            checkbox.click()
            self.log("☑️ Checkbox de concordância marcado")

            # botão Enviar (input submit)
            botao_enviar = WebDriverWait(self.driver, 1).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//input[@type='submit' and @name='Enviar']"
                ))
            )
            botao_enviar.click()
            self.log("📨 Formulário enviado")

            # aguarda popup fechar
            WebDriverWait(self.driver, 1).until(
                lambda d: len(d.window_handles) == 1
            )

            # volta para janela principal
            self.driver.switch_to.window(janela_principal)
            self.log("✅ Acesso às petições liberado com sucesso")

        except TimeoutException:
            self.log("🔓 Nenhum popup de amplo acesso detectado")
            self.driver.switch_to.window(janela_principal)
                
    def abrir_detalhe_processo(self):
        numero = self.numero_atual
        if not numero:
            self.ui_warning(self, "Aviso", "Informe o número do processo.")
            return

        try:
            self.log("🔐 Garantindo login...")
            self.garantir_login()
            self.tratar_popups_login()
            self.log("🔎 Acessando página de pesquisa...")
            self.driver.get(URL_DESTINO)

            # aguarda campo NumPedido visível
            campo = WebDriverWait(self.driver, 1).until(
                EC.visibility_of_element_located((By.NAME, "NumPedido"))
            )
            campo.clear()
            campo.send_keys(numero)

            # submit
            campo.submit()

            # aguarda link de detalhe
            self.log("📄 Aguardando link de detalhe...")
            link = WebDriverWait(self.driver, 1).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "a[href*='MarcasServletController?Action=detail']")
                )
            )

            # clica (não usa get!)
            link.click()
    
            # aguarda página de detalhe
            WebDriverWait(self.driver, 1).until(
    EC.presence_of_element_located((By.TAG_NAME, "body"))
)
            time.sleep(1)  # estabilidade
            print("🌐 URL atual:", self.driver.current_url)
           # self.extrair_dados_pagina()
            # 🔓 garante acesso às petições (se necessário)
            self.garantir_acesso_peticiones()
            self.liberar_acesso_peticiones()    # trata o popup
            # ✅ tenta baixar PDF
            self.tentar_clicar_botao_pdf(numero)
        
        except UnexpectedAlertPresentException:
            self.tratar_popups_login()
            self.log("⚠️ Alert inesperado tratado, retomando fluxo")
        except TimeoutException:
            if "Nenhum resultado foi encontrado" in self.driver.page_source:
                self.log(f"❌ Processo {numero} inexistente no INPI")
                self._registrar_processo_concluido(numero)
            else:
                self.log(f"⏱ Timeout inesperado ao abrir processo {numero}")

            QTimer.singleShot(0, self._finalizar_processo_atual)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao abrir detalhe do processo:")
            
            QTimer.singleShot(0, self.parar_processamento)
             
    def tentar_clicar_botao_pdf(self, numero):
          try:
              img = self.driver.find_element(
                  By.XPATH,
                  "//div[@id='389' or @id='394']/ancestor::tr//img[contains(@class,'salvaDocumento')]"
              )
  
              img.click()
              self.log("Clique no ícone PDF executado para ID 389 ou 394.")
  
              QTimer.singleShot(
                600,
                lambda: QTimer.singleShot(600, self.tratar_modal_captcha)

            )

  
          except Exception :
              self.log(f"Nenhum PDF encontrado para ID 389 ou 394")
              self._registrar_processo_concluido(numero)
              QTimer.singleShot(0, self._finalizar_processo_atual)
      
    
    #def download_iniciado(self):
    #    for f in os.listdir(DOWNLOAD_DIR):
    #        if f.lower().endswith(".crdownload"):
    #            return True
    #    return False
    def _repetir_processo_atual(self, motivo):
        numero = self.numero_atual
      
        self.tentativas_por_processo.setdefault(numero, 0)
        self.tentativas_por_processo[numero] += 1
      
        self.log(
            f"⚠️ Falha no download do processo {numero} "
            f"(tentativa {self.tentativas_por_processo[numero]}): {motivo}"
        )
      
        if self.tentativas_por_processo[numero] >= self.MAX_TENTATIVAS:
            self.log(f"❌ Processo {numero} excedeu tentativas e será ignorado.")
           # self._registrar_processo_concluido(numero)
            self.numero_atual = None
            self.processando = False
            QTimer.singleShot(0, self._processar_proximo)
            return
      
        # 🔁 REAGENDAR O MESMO PROCESSO
        self.numero_atual = None
        self.processando = False
      
        QTimer.singleShot(3000, self._processar_proximo)

    def tratar_modal_captcha(self):
        """
        Fluxo:
        - aguarda modal #janelaModalCaptchaDownload visível (ou .g-recaptcha)
        - tenta clicar checkbox do reCAPTCHA (dentro do iframe)
        - aguarda que o token apareça (no DOM: g-recaptcha-response ou input #recaptcha-token)
        - quando token presente, clica botão #captchaButton (Download)
        - aguarda download terminar e processa PDF
        """
        try:
            self.log("[captcha] aguardando modal...")
            # espera até o modal aparecer (ou timeout)
            t0 = time.time()
            modal = None
            while time.time() - t0 < WAIT_LONG:
                try:
                    modal = self.driver.find_element(By.CSS_SELECTOR, "#janelaModalCaptchaDownload")
                    if modal.is_displayed():
                        break
                except NoSuchElementException:
                    # fallback: procurar .g-recaptcha direto
                    try:
                        g = self.driver.find_element(By.CSS_SELECTOR, ".g-recaptcha")
                        if g.is_displayed():
                            break
                    except NoSuchElementException:
                        pass
                time.sleep(1)

            if not modal:
                self.log("[captcha] modal não encontrado explicitamente — tentando detectar .g-recaptcha")
            else:
                self.log("[captcha] modal visível")

            # tenta clicar a checkbox do reCAPTCHA (iframe com src contendo 'anchor')
            iframe = self.selenium_get_recaptcha_iframe()
            if iframe:
                try:
                    self.driver.switch_to.frame(iframe)
                    # checkbox id recaptcha-anchor
                    try:
                        checkbox = WebDriverWait(self.driver, WAIT_SHORT).until(
                            EC.element_to_be_clickable((By.ID, "recaptcha-anchor"))
                        )
                        checkbox.click()
                        self.log("[captcha] checkbox clicado")
                    except Exception as e:
                        self.log(f"[captcha] falha ao clicar checkbox: ")
                    finally:
                        self.driver.switch_to.default_content()
                except Exception as e:
                    self.log(f"[captcha] erro switch_to.frame: ")
            else:
                self.log("[captcha] iframe do recaptcha não encontrado")
            div = self.identificar_captcha_imagem()
            if div:
                try:
                    self.driver.switch_to.frame(div)
                    # checkbox id recaptcha-anchor
                    try:
                        
                        checkbox = WebDriverWait(self.driver, WAIT_SHORT).until(
                            EC.element_to_be_clickable((By.ID, "solver-button"))
                        )
                        checkbox.click()
                        self.log("[buster] checkbox clicado")
                    except Exception as e:
                        self.log(f"[buster] falha ao clicar checkbox: ")
                    finally:
                        self.driver.switch_to.default_content()
                except Exception as e:
                    self.log(f"[buster] erro switch_to.frame:")
            else:
                self.log("[buster] iframe do recaptcha não encontrado")

            # Se você usa Buster: a extensão pode interagir com o desafio automaticamente.
            # Aqui aguardamos o token aparecer no DOM (g-recaptcha-response ou input#recaptcha-token)
            self.log("[captcha] aguardando token resolver (sem timeout)...")
            token = None
            timeout = 420  # segundos
            inicio = time.time()

            while time.time() - inicio < timeout:
                try:
                    # 1) input hidden recaptcha-token (algumas implementações do INPI colocam o token aqui)
                   
                    try:
                        token_input = self.driver.find_element(By.ID, "recaptcha-token")
                        val = token_input.get_attribute("value")
                        if val and len(val) > 10:
                            token = val
                            break
                    except NoSuchElementException:
                        pass

                    # 2) textarea.g-recaptcha-response
                    try:
                        gr = self.driver.find_element(By.CSS_SELECTOR, "textarea.g-recaptcha-response")
                        val2 = gr.get_attribute("value")
                        if val2 and len(val2) > 10:
                            token = val2
                            break
                    except NoSuchElementException:
                        pass

                except Exception as e:
                    print("debug token check:")
                time.sleep(1)

            if token:
                self.log(f"[captcha] token detectado (len={len(token)})")
                # injeta token em possíveis campos e aciona o botão download
                try:
                    # tentar preencher g-recaptcha-response via JS (algumas páginas aceitam)
                    script_set = """
                    (function(t){
                        var ga = document.querySelector('textarea.g-recaptcha-response');
                        if(ga){ ga.value = t; ga.style.display='block'; }
                        var ri = document.getElementById('recaptcha-token');
                        if(ri) ri.value = t;
                        return true;
                    })(arguments[0]);
                    """
                    self.driver.execute_script(script_set, token)
                except Exception as e:
                    print("Erro ao injetar token via JS:")

                # clica no botão #captchaButton (Download)
                try:
                    btn = WebDriverWait(self.driver,1).until(
                        EC.element_to_be_clickable((By.ID, "captchaButton"))
                    )
                    btn.click()
                    self.log("[captcha] botão Download clicado")
                # aguarda o PDF ser baixado
                    caminho_pdf = self.selenium.wait_for_download()

                    if caminho_pdf:
                        self._renomear_pdf_para_processo(caminho_pdf)
                    else:
                        self.log("❌ PDF não foi baixado.")
                        QTimer.singleShot(0, self._finalizar_processo_atual)
    
                except Exception as e:
                    self.log(f"[captcha] falha ao clicar #captchaButton: {e}")
            else:
                self.log("[captcha] token não detectado — talvez Buster não resolveu automaticamente.")

           
                self.log("[captcha] nenhum PDF detectado no diretório de download (timeout).")
        except Exception as e:
            self.log(f"[captcha] erro no tratamento do modal: {e}")
                
   
    def _renomear_pdf_para_processo(self, caminho_pdf):
        try:
            numero = self.numero_atual
            if not numero:
                self.log("⚠️ Processo atual não definido para renomear PDF.")
                return

            # sanitiza o número (remove caracteres inválidos)
            nome_seguro = "".join(c for c in numero if c.isalnum())

            novo_nome = f"{nome_seguro}.pdf"
            novo_caminho = DOWNLOAD_DIR / novo_nome

            # se já existir, adiciona timestamp
            if novo_caminho.exists():
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                novo_caminho = DOWNLOAD_DIR / f"{nome_seguro}_{ts}.pdf"

            shutil.move(caminho_pdf, novo_caminho)

            self.log(f"📄 PDF renomeado para: {novo_caminho.name}")
            self._registrar_processo_concluido(numero)
            QTimer.singleShot(0, self._finalizar_processo_atual)
        except Exception as e:
            self.log(f"❌ Erro ao renomear PDF: {e}")

        

    def identificar_captcha_imagem(self) -> bool:
        """
        Retorna True se o CAPTCHA de seleção de imagens estiver presente.
        NÃO interage com o CAPTCHA.
        """
        try:
            # procura iframe do challenge (bframe)
            iframes = self.driver.find_elements(By.TAG_NAME, "div")
    
            for iframe in iframes:
                src = iframe.get_attribute("src") or ""
                if "api2/bframe" in src:
                    # entra no iframe do desafio
                    self.driver.switch_to.frame(iframe)

                    try:
                        # procura a div principal do desafio de imagens
                        self.driver.find_element(By.ID, "solver-button")
                        return True
                    except NoSuchElementException:
                        pass
                    finally:
                        # sempre volta para o DOM principal
                        self.driver.switch_to.default_content()

            return False

        except Exception:
            self.driver.switch_to.default_content()
            return False

    def selenium_get_recaptcha_iframe(self):
        # procura iframe que contém 'anchor' (checkbox) ou 'api2/anchor'
        try:
            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            for fr in iframes:
                src = fr.get_attribute("src") or ""
                if "api2/anchor" in src or "recaptcha" in src and "anchor" in src:
                    return fr
            # fallback: iframe title containing "reCAPTCHA"
            for fr in iframes:
                title = fr.get_attribute("title") or ""
                if "reCAPTCHA" in title or "recaptcha" in title.lower():
                    return fr
        except Exception:
            pass
        return None 
           
    def _pdf_baixado_com_sucesso(self, caminho_final):
        caminho_pdf = os.path.abspath(caminho_final)

        self.log(f"📥 PDF salvo: {caminho_pdf}")

        QTimer.singleShot(0, self._finalizar_processo_atual)

    
    def _atualizar_contador_ui(self):
        texto = f"Processos extraídos: {self.processos_extraidos} / {self.total_processos}"
        self.label_contador.setText(texto)
        
    def _finalizar_processo_atual(self):
        numero = self.numero_atual  # ✅ já existe

        #self._registrar_processo_concluido(numero)

        self.processos_extraidos += 1
        self._atualizar_contador_ui()

        self.log(f"✔ Processo {numero} finalizado")

        self.numero_atual = None
        self.processando = False

        QTimer.singleShot(200, self._processar_proximo)

                                     
                        
    def closeEvent(self, event):
        # encerra selenium quando fechar UI
        try:
            self.selenium.stop()
        except Exception:
            pass
        event.accept()


# Se você quer testar este arquivo standalone (sem main.py),
# comente a importação em main.py e execute este módulo diretamente.
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    w = MainApp()
    w.show()
    sys.exit(app.exec_())
