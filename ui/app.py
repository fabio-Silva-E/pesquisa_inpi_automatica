# ui/app.py
import pyautogui
import os
import time
import sys
import tempfile
import shutil
import requests
import threading
print(requests.get("https://api.ipify.org").text)

from collections import deque
from pathlib import Path
from urllib.parse import urljoin
from datetime import datetime
from openpyxl import Workbook, load_workbook


from PyQt5.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QPushButton, QLineEdit, QListWidget,
    QLabel, QMessageBox, QApplication, QProgressDialog, QTextEdit, QMessageBox, QComboBox
)
from PyQt5.QtCore import QTimer, QDateTime, QThread, pyqtSignal

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, WebDriverException,
    ElementClickInterceptedException, UnexpectedAlertPresentException,
    NoAlertPresentException
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# suas funções internas

from ui.alert_menssenger import mostrar_toast
from .vpn_manager import VPNManager


def pasta_app():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent.parent

BASE_DIR = pasta_app()

def resource_path(relative):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath("."), relative)

PROFILE_PATH = BASE_DIR / "ui"/ "chrome_profile"

EXCEL_PROCESSOS_PATH = BASE_DIR / "processos.xlsx"

caminho = BASE_DIR / "solver_button.png"

DOWNLOAD_DIR = BASE_DIR / "pdfs"
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

from openpyxl import Workbook

if not EXCEL_PROCESSOS_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Numero_Processo"
    wb.save(EXCEL_PROCESSOS_PATH)

# Chromedriver / Chrome options
HEADLESS = False  # não usar headless para extensões
CHROME_BINARY = None  # se usa caminho custom do navegador (opcional)

# Timeouts
WAIT_SHORT = 2
WAIT_MEDIUM = 8
WAIT_LONG = 25

# INPI URLs
URL_INPI = "https://busca.inpi.gov.br/pePI/"
URL_DESTINO = "https://busca.inpi.gov.br/pePI/jsp/marcas/Pesquisa_num_processo.jsp"

# -------------------------------------------

class SeleniumController:
    def __init__(self):
        self.driver = None

    def start(self):
        from selenium.webdriver.chrome.service import Service

        chrome_options = Options()
        
        # 🔐 worker_id dinâmico (não quebra código antigo)
        
        worker_id = getattr(self, "worker_id", 1)

        # 📁 PERFIL ÚNICO POR WORKER
        profile_path = os.path.join(PROFILE_PATH, f"profile_{worker_id}")
        os.makedirs(profile_path, exist_ok=True)
        chrome_options.add_argument(f"--user-data-dir={profile_path}")

        # 📁 DOWNLOAD ÚNICO POR WORKER
        download_dir = os.path.join(DOWNLOAD_DIR, f"worker_{worker_id}")
        os.makedirs(download_dir, exist_ok=True)
       
        prefs = {
            "download.default_directory": os.path.abspath(download_dir),
            "download.prompt_for_download": False,
            "plugins.always_open_pdf_externally": True,
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
        }
        chrome_options.add_experimental_option("prefs", prefs)

        # Flags que você já usa
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-features=PasswordLeakDetection")
        chrome_options.add_argument("--disable-save-password-bubble")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--autoplay-policy=no-user-gesture-required")
        chrome_options.add_argument("--verbose")
        chrome_options.add_argument("--no-first-run")
        chrome_options.add_argument("--disable-software-rasterizer")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")

        service = Service()
        service.start_timeout = 60

        self._aguardar_rede_estavel()

        try:
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.set_window_size(600, 720)

            # 🔥 guarda para uso futuro (downloads, logs, etc)
            self.download_dir = download_dir

            return self.driver
        except Exception as e:
            raise RuntimeError(f"Erro ao iniciar ChromeDriver: {e}")


    
    def _aguardar_rede_estavel(self, timeout=30):
        print("⏳ Aguardando rede estabilizar após a troca de VPN...")
        inicio = time.time()

        while time.time() - inicio < timeout:
            try:
                r = requests.get("https://www.google.com", timeout=5)
                if r.status_code == 200:
                    print("✅ Rede estabilizada.")
                    return
            except requests.exceptions.RequestException:
                pass
            time.sleep(3)  # Aguardar mais tempo
        raise RuntimeError("❌ A rede não estabilizou após a troca de VPN.")

    def stop(self):
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass

    def wait_for_element(self, by, value, timeout=WAIT_MEDIUM):
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))

    def wait_for_clickable(self, by, value, timeout=WAIT_MEDIUM):
        return WebDriverWait(self.driver, timeout).until(EC.element_to_be_clickable((by, value)))

    def safe_click(self, by, value, timeout=WAIT_MEDIUM):
        try:
            el = self.wait_for_clickable(by, value, timeout=timeout)
            el.click()
            return True
        except Exception as e:
            print(f"[selenium] safe_click falhou ({by},{value}): {e}")
            return False

    def element_exists(self, by, value):
        try:
            self.driver.find_element(by, value)
            return True
        except NoSuchElementException:
            return Falsez

    def get_iframe_by_src_contains(self, partial_src):
        iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
        for fr in iframes:
            try:
                src = fr.get_attribute("src") or ""
                if partial_src in src:
                    return fr
            except Exception:
                continue
        return None
    
    def popup_erro_download_visivel(self):
        try:
            popup = self.driver.find_element(
                By.XPATH,
                "//div[contains(text(),'Erro') or contains(text(),'erro')]"
            )
            return popup.is_displayed()
        except:
            return False
        
   
    def wait_for_download(self, worker_id, timeout=90):
        """
        Aguarda o ÚLTIMO PDF realmente baixado (por data de modificação)
        """
        t0 = time.time()
      
        # 🔥 pasta exclusiva do worker
        download_dir = DOWNLOAD_DIR / f"worker_{worker_id}"
        download_dir.mkdir(parents=True, exist_ok=True)
      
        # snapshot inicial
        arquivos_iniciais = {
            f: os.path.getmtime(download_dir / f)
            for f in os.listdir(download_dir)
            if f.lower().endswith(".pdf")
        }
      
        while time.time() - t0 < timeout:
            if self.popup_erro_download_visivel():
                print("❌ Erro exibido pelo site durante download")
                raise RuntimeError("❌ Erro exibido pelo site durante download")
      
            pdfs = []
      
            for nome in os.listdir(download_dir):
                if not nome.lower().endswith(".pdf"):
                    continue
      
                caminho = download_dir / nome
      
                try:
                    mtime = os.path.getmtime(caminho)
                except FileNotFoundError:
                    continue
      
                # só PDFs criados/modificados depois do clique
                if nome not in arquivos_iniciais or mtime > arquivos_iniciais.get(nome, 0):
                    if self._arquivo_estavel(caminho):
                        pdfs.append((mtime, caminho))
      
            if pdfs:
                pdfs.sort(key=lambda x: x[0], reverse=True)
                return pdfs[0][1]
      
            time.sleep(0.5)
      
        return None

    def _arquivo_estavel(self, caminho, tentativas=5, intervalo=0.5):
        """
        Verifica se o tamanho do arquivo não muda (download finalizado)
        """
        try:
            tamanho_anterior = -1
            for _ in range(tentativas):
                tamanho = os.path.getsize(caminho)
                if tamanho == tamanho_anterior:
                    return True
                tamanho_anterior = tamanho
                time.sleep(intervalo)
        except FileNotFoundError:
            return False
        return False   
    
class SeleniumWorker(QThread):
    finished = pyqtSignal(int)
    error = pyqtSignal(int, str)

    def __init__(self, app, worker_id):
        super().__init__()
        self.app = app
        self.worker_id = worker_id

    def run(self):
        try:
            if self.worker_id == 1:
                driver = self.app.driver1
            elif self.worker_id == 2:
                driver = self.app.driver2
            elif self.worker_id == 3:
                driver = self.app.driver3
            else:
                raise ValueError("Worker inválido")

            self.app.abrir_detalhe_processo(driver)
            self.finished.emit(self.worker_id)

        except Exception as e:
            self.error.emit(self.worker_id, str(e))



      
class MainApp(QWidget):
    log_signal = pyqtSignal(str)
    def __init__(self):
        super().__init__()
        self.log_signal.connect(self._log_ui)

        self.tentativas_por_processo = {}
        self.MAX_TENTATIVAS = 2
        
        # diretórios de download por worker
        self.download_dirs = {}
        
        # 1️⃣ Flags e dados
        self.parar_loop = False
        self.loop_ativo = False
        self.processando = {1: False, 2: False}

        self.numero_atual = {
            1: None,
            2: None
        }

        
        self.indice_processo = 0
        self.total_processos = 0
        self.processos_extraidos = 0
        
        self.arquivo_concluidos = "processos_concluidos.txt"
        self.processos = []
        self.processos_concluidos = set()
        
        self.vpn = VPNManager()
      
       
       
        # 2️⃣ Criar UI PRIMEIRO
        self._criar_interface()
        self._iniciar_monitor_ip()
        # 3️⃣ Carregar dados DEPOIS
        self._carregar_processos_concluidos()
        self.atualizar_lista_processos()

        # 4️⃣ Filtrar concluídos
        self.processos = [
            p for p in self.processos
            if p not in self.processos_concluidos
        ]
        for wid in (1, 2, 3):
           pasta = Path(DOWNLOAD_DIR) / f"worker_{wid}"
           pasta.mkdir(parents=True, exist_ok=True)
           self.download_dirs[wid] = pasta
        self.carregar_usuarios_excel()
         
        for u in self.usuarios:
            self.combo_usuario.addItem(u["usuario"])
            self.combo_usuario_2.addItem(u["usuario"])
            self.combo_usuario_3.addItem(u["usuario"])
            
        # 🔌 CONECTA OS COMBOS AOS MÉTODOS
        self.combo_usuario.currentIndexChanged.connect(self.on_usuario_selecionado)
        self.combo_usuario_2.currentIndexChanged.connect(self.on_usuario_selecionado_2)
        self.combo_usuario_3.currentIndexChanged.connect(self.on_usuario_selecionado_3)

    def _criar_interface(self):
        self.setWindowTitle("INPI - vs(1.2) - 02/02/2026")
        self.setGeometry(50, 30, 700, 50)
      
        # 🔹 Layout principal
        self.layout = QHBoxLayout(self)
      
        # ==================================================
        # 🟦 COLUNA LOGIN
        # ==================================================
        coluna_login = QVBoxLayout()
      
        self.label_ip = QLabel("🌍 IP: ---")
        coluna_login.addWidget(self.label_ip)
        # ===== dropdown de timeout =====
        coluna_login.addWidget(QLabel("Timeout Captcha (segundos)"))

        self.combo_timeout = QComboBox()
        self.combo_timeout.addItem("Selecione o timeout")
        self.combo_timeout.model().item(0).setEnabled(False)
        
        # opções
        self.combo_timeout.addItems(["30", "60", "120", "180", "300"])
        
        # valor padrão
        self.combo_timeout.setCurrentText("120")
        
        coluna_login.addWidget(self.combo_timeout)

        
        # ===== USUÁRIO 1 =====
        coluna_login.addWidget(QLabel("Login Aba 1"))
        self.combo_usuario = QComboBox()
        self.combo_usuario.addItem("Selecione o usuário INPI")
        self.combo_usuario.model().item(0).setEnabled(False)
      
        self.input_usuario = QLineEdit()
        self.input_usuario.setReadOnly(True)
      
        self.input_senha = QLineEdit()
        self.input_senha.setEchoMode(QLineEdit.Password)
        self.input_senha.setReadOnly(True)
      
        coluna_login.addWidget(self.combo_usuario)
        coluna_login.addWidget(self.input_usuario)
        coluna_login.addWidget(self.input_senha)
      
        # ===== USUÁRIO 2 =====
        coluna_login.addWidget(QLabel("Login Aba 2"))
        self.combo_usuario_2 = QComboBox()
        self.combo_usuario_2.addItem("Selecione o usuário INPI")
        self.combo_usuario_2.model().item(0).setEnabled(False)
      
        self.input_usuario_2 = QLineEdit()
        self.input_senha_2 = QLineEdit()
        self.input_senha_2.setEchoMode(QLineEdit.Password)
      
        coluna_login.addWidget(self.combo_usuario_2)
        coluna_login.addWidget(self.input_usuario_2)
        coluna_login.addWidget(self.input_senha_2)
      
        # ===== USUÁRIO 3 =====
        coluna_login.addWidget(QLabel("Login Aba 3"))
        self.combo_usuario_3 = QComboBox()
        self.combo_usuario_3.addItem("Selecione o usuário INPI")
        self.combo_usuario_3.model().item(0).setEnabled(False)
      
        self.input_usuario_3 = QLineEdit()
        self.input_senha_3 = QLineEdit()
        self.input_senha_3.setEchoMode(QLineEdit.Password)
      
        coluna_login.addWidget(self.combo_usuario_3)
        coluna_login.addWidget(self.input_usuario_3)
        coluna_login.addWidget(self.input_senha_3)
      
        self.btn_iniciar = QPushButton("🌐 Abrir site INPI")
        self.btn_iniciar.clicked.connect(self.iniciar_selenium)
        coluna_login.addWidget(self.btn_iniciar)
      
        coluna_login.addStretch()
      
        # ==================================================
        # 🟨 COLUNA PROCESSOS
        # ==================================================
        coluna_processos = QVBoxLayout()
      
        coluna_processos.addWidget(QLabel("📄 Processos"))
      
        self.lista_processos = QListWidget()
        self.lista_processos.itemClicked.connect(self.selecionar_processo)
        coluna_processos.addWidget(self.lista_processos)
      
        self.entry_processo = QLineEdit()
        self.entry_processo.setPlaceholderText("Número do processo")
        coluna_processos.addWidget(self.entry_processo)
      
        self.btn_buscar = QPushButton("▶ Iniciar extração")
        self.btn_buscar.clicked.connect(self.iniciar_processamento_em_lote)
      
        self.btn_parar = QPushButton("⏹ Parar")
        self.btn_parar.clicked.connect(self.parar_processamento)
      
        coluna_processos.addWidget(self.btn_buscar)
        coluna_processos.addWidget(self.btn_parar)
      
        self.label_status = QLabel("Status: idle")
        coluna_processos.addWidget(self.label_status)
      
        self.label_contador = QLabel("Processos extraídos: 0 / 0")
        self.label_contador.setStyleSheet("font-size: 11pt; font-weight: bold;")
        coluna_processos.addWidget(self.label_contador)
      
        coluna_processos.addStretch()
      
        # ==================================================
        # 🟩 COLUNA LOGS
        # ==================================================
        coluna_logs = QVBoxLayout()
      
        coluna_logs.addWidget(QLabel("🧾 Logs"))
      
        self.console_log = QTextEdit()
        self.console_log.setReadOnly(True)
        self.console_log.setStyleSheet("""
            background-color: white;
            color: black;
            font-family: Consolas;
            font-size: 10pt;
        """)
        coluna_logs.addWidget(self.console_log)
      
        # ==================================================
        # 🔗 ADICIONA AO LAYOUT PRINCIPAL (AGORA SIM)
        # ==================================================
        self.layout.addLayout(coluna_login, 2)
        self.layout.addLayout(coluna_processos, 2)
        self.layout.addLayout(coluna_logs, 3)

    def _log_ui(self, mensagem):
        timestamp = QDateTime.currentDateTime().toString("HH:mm:ss")
        self.console_log.append(f"[{timestamp}] {mensagem}")
        self.console_log.ensureCursorVisible()
        self.label_status.setText(f"Status: {mensagem}")

    def _iniciar_monitor_ip(self):
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.atualizar_ip)
        self.timer.start(5000)
        self.atualizar_ip()
    
    def atualizar_ip(self):
        ip = self.vpn.ip_atual()
        self.label_ip.setText(f"🌍 IP: {ip}")

    def trocar_vpn(self):
        print("BOTÃO CLICADO")
        self.label_ip.setText("🔄 Trocando VPN...")
        ip = self.vpn.conectar()
        self.label_ip.setText(f"🌍 IP: {ip}")
        
    def on_usuario_selecionado_2(self, index):
        if index <= 0:
            self.input_usuario_2.clear()
            self.input_senha_2.clear()
            return
        user = self.usuarios[index - 1]
        self.input_usuario_2.setText(user["usuario"])
        self.input_senha_2.setText(user["senha"])
        
    def on_usuario_selecionado_3(self, index):
        if index <= 0:
            self.input_usuario_3.clear()
            self.input_senha_3.clear()
            return
        user = self.usuarios[index - 1]
        self.input_usuario_3.setText(user["usuario"])
        self.input_senha_3.setText(user["senha"])


    def on_usuario_selecionado(self, index):
        if index == 0:
            self.input_usuario.clear()
            self.input_senha.clear()
            return
    
        user = self.usuarios[index - 1]  # 👈 deslocamento
        self.input_usuario.setText(user["usuario"])
        self.input_senha.setText(user["senha"])

   
    def parar_processamento(self):
        if not self.loop_ativo:
            return

        self.loop_ativo = False
        self.processando = False

        self.log("🛑 Processamento interrompido pelo usuário.")
        self.label_status.setText("Status: parado")
        self.lista_processos.setEnabled(True)
   
    def _carregar_processos_concluidos(self):
        if not os.path.exists(self.arquivo_concluidos):
            return

        with open(self.arquivo_concluidos, "r", encoding="utf-8") as f:
            for linha in f:
                numero = linha.strip()
                if numero:
                    self.processos_concluidos.add(numero)
    
           

    
       
    from collections import deque

    def iniciar_processamento_em_lote(self):
    
        # 🔐 Verifica se os Chromes estão ativos
        if not all(hasattr(self, f"driver{i}") for i in (1, 2, 3)):
            self.ui_warning(
                "Chrome não iniciado",
                "Abra os três Chromes antes de iniciar a extração."
            )
            self.log("⚠️ Tentativa de iniciar lote sem Chromes abertos.")
            return
    
        if not self.processos:
            self.ui_warning("Aviso", "Nenhum processo carregado.")
            return
    
        self.loop_ativo = True
        self.lista_processos.setEnabled(False)
    
        # 🔥 CONVERTE LISTA EM 3 FILAS
        self.processos_1 = deque()
        self.processos_2 = deque()
        self.processos_3 = deque()
    
        for i, p in enumerate(self.processos):
            if i % 3 == 0:
                self.processos_1.append(p)
            elif i % 3 == 1:
                self.processos_2.append(p)
            else:
                self.processos_3.append(p)
    
        self.total_processos = len(self.processos)
        self.processos_extraidos = 0
        self._atualizar_contador_ui()
    
        self.log("🚀 Iniciando processamento em lote (3 Chromes)...")
    
        # 🚀 DISPARA OS 3 WORKERS
        self._processar_proximo(1)
        self._processar_proximo(2)
        self._processar_proximo(3)

    def clicar_imagem(
            self,
            timeout=15,
            confidence=0.85,
            clicar=True,
            delay=0.5
        ):
            """
            Localiza uma imagem PNG na tela e opcionalmente clica nela.
        
            :param nome_imagem: Nome do arquivo PNG (ex: 'captcha_checkbox.png')
            :param timeout: Tempo máximo de espera (segundos)
            :param confidence: Precisão da imagem (0.7 a 0.95)
            :param clicar: Se True, clica no centro da imagem
            :param delay: Delay após clicar
            :return: (x, y) ou None
            """
        
            caminho = BASE_DIR / "solver_button.png"
            if not caminho.exists():
                raise FileNotFoundError(f"Imagem não encontrada: {caminho}")
        
            inicio = time.time()
        
            while time.time() - inicio < timeout:
                pos = pyautogui.locateCenterOnScreen(
                    str(caminho),
                    confidence=confidence
                )
        
                if pos:
                    if clicar:
                        pyautogui.moveTo(pos.x, pos.y, duration=0.2)
                        pyautogui.click()
                        time.sleep(delay)
                    return pos
        
                time.sleep(0.3)
            self.log(f"⚠️ Imagem não encontrada na tela: {nome_imagem}")
            return None             

    def carregar_usuarios_excel(self):
        self.usuarios = []  # lista de dicts

        caminho = BASE_DIR / "credenciais.xlsx"
        if not caminho.exists():
            self.log("⚠️ Arquivo credenciais.xlsx não encontrado.")
            return

        wb = load_workbook(caminho, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue

            self.usuarios.append({
                "usuario": str(row[0]).strip(),
                "senha": str(row[1]).strip()
            })
    
    def _erro_worker(self, worker_id, mensagem):
        self.log(f"❌ Worker {worker_id} — Erro: {mensagem}")
        self.processando[worker_id] = False
    
        QTimer.singleShot(
            0, lambda: self._processar_proximo(worker_id)
        )


    def _processar_proximo(self, worker_id):
        # 🚫 Evita rodar dois processos ao mesmo tempo no mesmo worker
        if self.processando.get(worker_id):
            return

        # 🔥 Seleciona a fila correta
        if worker_id == 1:
            fila = self.processos_1
        elif worker_id == 2:
            fila = self.processos_2
        elif worker_id == 3:
            fila = self.processos_3
        else:
            return

        # 🏁 Fila vazia
        if not fila:
            self.log(f"🏁 Fila finalizada (Worker {worker_id})")
            return

        self.processando[worker_id] = True

        numero = fila.popleft()
        self.numero_atual[worker_id] = numero
        self.tentativas_por_processo.setdefault(numero, 0)

        self.log(
            f"📦 Tela {worker_id} processando {numero} "
            f"(tentativa {self.tentativas_por_processo[numero] + 1})"
        )

        worker = SeleniumWorker(self, worker_id)
        worker.finished.connect(self._finalizar_processo_atual)
        worker.error.connect(self._erro_worker)
        worker.start()

        # 🔒 mantém referência (IMPORTANTE!)
        setattr(self, f"_worker_{worker_id}", worker)


    def selecionar_processo(self, item):
        if self.processando:
            return  # ignora clique durante processamento em lote

        numero = item.text()
        self.entry_processo.setText(numero)
        self.log(f"📌 Processo selecionado: {numero}")
   
    def _registrar_processo_concluido(self, numero):
        if not numero:
            self.log("⚠️ Tentativa de registrar processo concluído com número inválido.")
            return

        with open(self.arquivo_concluidos, "a", encoding="utf-8") as f:
            f.write(str(numero) + "\n")

        self.processos_concluidos.add(numero)

  
    def ui_warning(self, titulo, msg):
        QTimer.singleShot(
            0,
            lambda: QMessageBox.warning(self, titulo, msg)
        )

    def ui_error(self, titulo, msg):
        QTimer.singleShot(
            0,
            lambda: QMessageBox.critical(self, titulo, msg)
        )
    
    def log(self, mensagem):
        self.log_signal.emit(mensagem)


    def atualizar_lista_processos(self):
        self.lista_processos.clear()
        self.processos = []

        try:
            if not EXCEL_PROCESSOS_PATH.exists():
                raise FileNotFoundError(
                    f"Arquivo não encontrado: {EXCEL_PROCESSOS_PATH}"
                )

            wb = load_workbook(EXCEL_PROCESSOS_PATH, data_only=True)
            ws = wb.active

            header = ws["A1"].value
            if header != "Numero_Processo":
                raise ValueError(
                    "A coluna A deve se chamar 'Numero_Processo'"
                )

            vistos = set()

            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if not row or not row[0]:
                    continue

                numero = str(row[0]).strip()

                # ignora duplicados e concluídos
                if numero in vistos or numero in self.processos_concluidos:
                    continue
    
                vistos.add(numero)
                self.processos.append(numero)
                self.lista_processos.addItem(numero)

            self.total_processos = len(self.processos)
            self._atualizar_contador_ui()

            self.log(f"📄 {len(self.processos)} processos carregados do Excel.")
    
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erro Excel",
                f"Falha ao carregar processos do Excel"
            )
            self.log(f"Falha ao carregar processos do Excel")

    def iniciar_selenium(self):
        self.log("Iniciando 2 Chromes independentes...")
   
        try:
            # Fecha se já existir
            if hasattr(self, "driver1") and self.driver1:
                self.selenium1.stop()
   
            if hasattr(self, "driver2") and self.driver2:
                self.selenium2.stop()
                
            if hasattr(self, "driver3") and self.driver2:
                self.selenium3.stop()
   
            time.sleep(3)
   
            # 🔹 DRIVER 1
            self.selenium1 = SeleniumController()
            self.selenium1.worker_id = 1
            self.driver1 = self.selenium1.start()
            self.driver1.get(URL_INPI)
   
            # 🔹 DRIVER 2
            self.selenium2 = SeleniumController()
            self.selenium2.worker_id = 2
            self.driver2 = self.selenium2.start()
            self.driver2.get(URL_INPI)
   
           # 🔹 DRIVER 3
            self.selenium3 = SeleniumController()
            self.selenium3.worker_id = 3
            self.driver3 = self.selenium3.start()
            self.driver3.get(URL_INPI)

            self.log("✅ Três Chromes iniciados com sucesso")
            self.iniciar_solver_auto()
            self._ultimo_solver_click = 0

        except Exception as e:
           QMessageBox.critical(self, "Erro", str(e))
    def login_duas_abas(self):
        self.login_inpi(self.driver1, self.input_usuario, self.input_senha)
        self.garantir_login(self.driver1)
        
        self.login_inpi(self.driver2, self.input_usuario_2, self.input_senha_2)
        self.garantir_login(self.driver2)

        self.log("✅ Login realizado nos dois navegadores")

    def _usuario_atual(self, worker_id):
        if worker_id == 1:
            return self.input_usuario
        elif worker_id == 2:
            return self.input_usuario_2
        elif worker_id == 3:
            return self.input_usuario_3
        return None


    def _senha_atual(self, worker_id):
        if worker_id == 1:
            return self.input_senha
        elif worker_id == 2:
            return self.input_senha_2
        elif worker_id == 3:
            return self.input_senha_3
        return None


    def login_inpi(self, driver, usuario_input, senha_input):
        driver.get("https://busca.inpi.gov.br/pePI/")
        time.sleep(3)

        driver.find_element(By.NAME, "T_Login").clear()
        driver.find_element(By.NAME, "T_Login").send_keys(usuario_input.text())

        driver.find_element(By.NAME, "T_Senha").clear()
        driver.find_element(By.NAME, "T_Senha").send_keys(senha_input.text())

        driver.find_element(
            By.XPATH,
            "//input[@type='submit' and contains(@value,'Continuar')]"
        ).click()
        
    def garantir_login(self, driver):

        """
         Garante que o usuário esteja:
        1) Logado
        2) Na página de pesquisa por número de processo (NumPedido visível)
        """

        try:
             # 🔎 Já está na página correta?
            driver.find_element(By.NAME, "NumPedido")
            self.log("✅ Usuário já está logado e na página correta.")
            return
        except:
            pass

        self.log("🔐 Usuário não está na página correta. Garantindo login...")
       
        # ======================
        # LOGIN
        # ======================
        try:
            worker_id = (
                1 if driver == self.driver1 else
                2 if driver == self.driver2 else
                3
            )
            
            self.login_inpi(
                driver,
                self._usuario_atual(worker_id),
                self._senha_atual(worker_id)
            )

        except Exception as e:
            raise Exception(f"Falha no login: {e}")

        # ======================
        # AGUARDA PÁGINA PRINCIPAL
         # ======================
        WebDriverWait(driver, WAIT_MEDIUM).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        self.log("🔑 Login realizado. Acessando menu Marcas...")

           # ======================
        # CLICA NO MENU MARCAS
        # ======================
        try:
            menu_marcas = WebDriverWait(driver, WAIT_MEDIUM).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "area[href*='Pesquisa_num_processo.jsp']")
                 )
            )
       
            # ⚠️ area precisa ser clicada via JS
            driver.execute_script("arguments[0].click();", menu_marcas)
       
        except TimeoutException:
            raise Exception("❌ Não foi possível localizar o menu Marcas")

        # ======================
        # AGUARDA CAMPO NumPedido
        # ======================
        try:
            WebDriverWait(driver, WAIT_MEDIUM).until(
                EC.presence_of_element_located((By.NAME, "NumPedido"))
            )
            self.log("✅ Página de pesquisa por processo carregada com sucesso.")
        except TimeoutException:
            raise Exception("❌ Campo NumPedido não apareceu após acessar Marcas")
    
    
   

    def ui_toast(self, mensagem, tempo=2000):
        QTimer.singleShot(
            0,
            lambda m=mensagem, t=tempo: mostrar_toast(m, t)
        )
   

    def possui_servico_389_ou_394(self, driver) -> bool:
        """
        Verifica se existe serviço 389 ou 394 na tabela de PDFs
        """
        try:
            driver.find_element(
                By.XPATH,
                "//a[normalize-space()='389' or normalize-space()='394']"
            )
            self.log("📄 Serviço 389 ou 394 detectado")
            return True
        except NoSuchElementException:
            self.log("📄 Serviço 389/394 NÃO encontrado")
           
            return False

    def garantir_acesso_peticiones(self, driver):
        
          

        try:
            
            self.log("🔍 Verificando necessidade de amplo acesso às petições...")
    
            link_amplo = WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//a[contains(normalize-space(),'Clique aqui para ter acesso as petições')]"
                ))
            )

            self.log("🔐 Acesso restrito detectado. Solicitando liberação...")
    
            driver.execute_script("arguments[0].scrollIntoView(true);", link_amplo)
            time.sleep(0.5)
    
            link_amplo.click()

            # aguarda o modal abrir (título da página do popup)
            WebDriverWait(driver, 1).until(
                EC.title_contains("Finalidade do Acesso")
            )

            self.log("🪟 Modal de acesso aberto")
    
            # dependendo do INPI, basta abrir o modal para liberar
            # aguarda voltar para a página principal
            WebDriverWait(driver, 1).until(
                EC.not_(EC.title_contains("Finalidade do Acesso"))
            )

            self.log("✅ Acesso às petições liberado")

        except TimeoutException:
            # não existe bloqueio → segue normal
            self.log("🔓 Nenhum bloqueio de petições detectado")
    
    def liberar_acesso_peticiones(self, driver):
        janela_principal = driver.current_window_handle

        try:
            self.log("🔍 Verificando popup de amplo acesso...")

            # aguarda abrir nova janela (popup)
            WebDriverWait(driver, 1).until(
                lambda d: len(d.window_handles) > 1
            )

            # muda para o popup
            for janela in driver.window_handles:
                if janela != janela_principal:
                    driver.switch_to.window(janela)
                    break

            self.log("🪟 Popup de amplo acesso detectado")

            # aguarda checkbox aparecer
            checkbox = WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable((By.ID, "aceite"))
            )
            checkbox.click()
            self.log("☑️ Checkbox de concordância marcado")

            # botão Enviar (input submit)
            botao_enviar = WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//input[@type='submit' and @name='Enviar']"
                ))
            )
            botao_enviar.click()
            self.log("📨 Formulário enviado")

            # aguarda popup fechar
            WebDriverWait(driver, 1).until(
                lambda d: len(d.window_handles) == 1
            )

            # volta para janela principal
            driver.switch_to.window(janela_principal)
            self.log("✅ Acesso às petições liberado com sucesso")

        except TimeoutException:
            self.log("🔓 Nenhum popup de amplo acesso detectado")
            driver.switch_to.window(janela_principal)
    
    
        
    def abrir_detalhe_processo(self, driver):
        if driver == self.driver1:
            worker_id = 1
        elif driver == self.driver2:
            worker_id = 2
        elif driver == self.driver3:
            worker_id = 3
        else:
            self.log("❌ Driver não reconhecido")
            return


        if not worker_id:
            self.log("❌ Driver não reconhecido")
            return

        numero = self.numero_atual.get(worker_id)

        if not numero:
            self.ui_warning("Aviso", "Informe o número do processo.")
            return

        try:
            self.log("🔐 Garantindo login...")
            self.garantir_login(driver)

            self.log("🔎 Acessando página de pesquisa...")
            driver.get(URL_DESTINO)

            campo = WebDriverWait(driver, 1).until(
                EC.visibility_of_element_located((By.NAME, "NumPedido"))
            )
            campo.clear()
            campo.send_keys(numero)
            campo.submit()

            self.log("📄 Aguardando link de detalhe...")
            link = WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "a[href*='MarcasServletController?Action=detail']")
                )
            )
            link.click()

            WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(1)

            self.log(f"🌐 URL atual (Tela {worker_id}): {driver.current_url}")
            # 🔎 Verifica se o serviço exige liberação de petições
            if not self.possui_servico_389_ou_394(driver):
                self.log("⏭️ Processo não exige liberação de petições")
            
                self._registrar_processo_concluido(numero)
                self._finalizar_processo_atual(worker_id)
            
                return

            self.garantir_acesso_peticiones(driver)
            self.liberar_acesso_peticiones(driver)

            # ✅ tenta baixar PDF
            self.tentar_clicar_botao_pdf(driver, worker_id, numero)

        except UnexpectedAlertPresentException:
            self.log("⚠️ Alert inesperado tratado, retomando fluxo")
     
        except TimeoutException:
            if "Nenhum resultado foi encontrado" in driver.page_source:
                self.log(f"❌ Processo {numero} inexistente no INPI")
                #self._registrar_processo_concluido(numero)
            else:
                self.log(f"⏱ Timeout inesperado ao abrir processo {numero}")
     
            QTimer.singleShot(0, lambda: self._finalizar_processo_atual(worker_id))
     
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erro",
                f"Falha ao abrir detalhe do processo:\n{e}"
            )
            QTimer.singleShot(0, self.parar_processamento)

    def tentar_clicar_botao_pdf(self, driver, worker_id, numero):
        try:
            elementos = driver.find_elements(
                By.XPATH,
                "//div[@id='389' or @id='394']/ancestor::tr//img[contains(@class,'salvaDocumento')]"
            )

            # ❌ IDs 389 e 394 NÃO existem
            if not elementos:
                self.log("Nenhum PDF encontrado para ID 389 ou 394")
                #self._registrar_processo_concluido(numero)
                QTimer.singleShot(0, lambda: self._finalizar_processo_atual(worker_id))
                return

            # ✅ Existe PDF → clica
            elementos[0].click()
            self.log("Clique no ícone PDF executado para ID 389 ou 394.")

            time.sleep(0.6)
            self.tratar_modal_captcha(driver, worker_id)

        except Exception as e:
            # erro REAL (selenium, timeout, DOM, etc)
            self.log(f"Erro ao tentar clicar no PDF: {e}")

    
    def _repetir_processo_atual(self, worker_id, motivo):
        numero = self.numero_atual[worker_id]

        self.tentativas_por_processo.setdefault(numero, 0)
        self.tentativas_por_processo[numero] += 1

        self.log(
            f"⚠️ Worker {worker_id} — Falha no download do processo {numero} "
            f"(tentativa {self.tentativas_por_processo[numero]}): {motivo}"
        )

        if self.tentativas_por_processo[numero] >= self.MAX_TENTATIVAS:
            self.log(
                f"❌ Worker {worker_id} — Processo {numero} excedeu tentativas e será ignorado."
            )

            # ❌ NÃO REGISTRA CONCLUÍDO AQUI (mantido como você tinha)
            self.numero_atual[worker_id] = None
            self.processando[worker_id] = False

            QTimer.singleShot(
                0, lambda: self._processar_proximo(worker_id)
            )
            return

        # 🔁 REAGENDAR O MESMO WORKER
        self.numero_atual[worker_id] = None
        self.processando[worker_id] = False

        QTimer.singleShot(
            3000, lambda: self._processar_proximo(worker_id)
        )
        
    def iniciar_solver_auto(self):
        if getattr(self, "_solver_ativo", False):
            self.log("⚠️ Solver automático já está rodando")
            return
    
        self._solver_ativo = True
    
        self._solver_thread = threading.Thread(
            target=self._loop_solver_button,
            daemon=True
        )
        self._solver_thread.start()
    
        self.log("🤖 Solver automático iniciado")

    
    def _loop_solver_button(self):
        self._ultimo_solver_click = 0  # inicializa UMA vez
    
        while self._solver_ativo:
            try:
                # 👇 CONTROLE DE INTERVALO REAL
                if time.time() - self._ultimo_solver_click > 1:
                    if self.clicar_solver_button():
                        self._ultimo_solver_click = time.time()
    
            except Exception as e:
                self.log(f"⚠️ Solver erro: {e}")
    
            time.sleep(1)  # 👈 loop roda a cada 1s
  
  
    def parar_solver_auto(self):
        self._solver_ativo = False
        self.log("🛑 Solver automático parado")

    def clicar_solver_button(self):

        time.sleep(2)

        img_path = BASE_DIR / "solver_button.png"

        if not img_path.exists():
            self.log("❌ solver_button.png não encontrado")
            return False

        try:
            pos = pyautogui.locateCenterOnScreen(
                str(img_path),   # 👈 CONVERSÃO OBRIGATÓRIA
                confidence=0.78
            )

            if pos:
                pyautogui.moveTo(pos.x, pos.y, duration=0.3)
                pyautogui.click()
                self.log("🤖 Solver button clicado via PyAutoGUI")
                return True
            else:
                self.log("⚠️ Solver button não encontrado na tela")
                return False

        except Exception as e:
            self.log(f"❌ Erro PyAutoGUI: {e}")
            return False

    def _selenium_por_worker(self, worker_id):
        if worker_id == 1:
            return self.selenium1
        elif worker_id == 2:
            return self.selenium2
        elif worker_id == 3:
            return self.selenium3
        else:
            raise ValueError("Worker inválido")
   
    def obter_timeout(self) -> int:
        try:
            return int(self.combo_timeout.currentText())
        except ValueError:
            return 120  # fallback seguro

    
    def tratar_modal_captcha(self, driver, worker_id):

        """
        Fluxo:
        - aguarda modal #janelaModalCaptchaDownload visível (ou .g-recaptcha)
        - tenta clicar checkbox do reCAPTCHA (dentro do iframe)
        - aguarda que o token apareça (no DOM: g-recaptcha-response ou input #recaptcha-token)
        - quando token presente, clica botão #captchaButton (Download)
        - aguarda download terminar e processa PDF
        """
        try:
            self.log("[captcha] aguardando modal...")
            # espera até o modal aparecer (ou timeout)
            t0 = time.time()
            modal = None
            while time.time() - t0 < WAIT_LONG:
                try:
                    modal = driver.find_element(By.CSS_SELECTOR, "#janelaModalCaptchaDownload")
                    if modal.is_displayed():
                        break
                except NoSuchElementException:
                    # fallback: procurar .g-recaptcha direto
                    try:
                        g = driver.find_element(By.CSS_SELECTOR, ".g-recaptcha")
                        if g.is_displayed():
                            break
                    except NoSuchElementException:
                        pass
                time.sleep(1)

            if not modal:
                self.log("[captcha] modal não encontrado explicitamente — tentando detectar .g-recaptcha")
            else:
                self.log("[captcha] modal visível")

            # tenta clicar a checkbox do reCAPTCHA (iframe com src contendo 'anchor')
            iframe = self.selenium_get_recaptcha_iframe(driver)
            if iframe:
                try:
                    driver.switch_to.frame(iframe)
                    # checkbox id recaptcha-anchor
                    try:
                        checkbox = WebDriverWait(driver, WAIT_SHORT).until(
                            EC.element_to_be_clickable((By.ID, "recaptcha-anchor"))
                        )
                        checkbox.click()
                        self.log("[captcha] checkbox clicado")
                        
                  
                    
                    except Exception as e:
                        self.log(f"[captcha] falha ao clicar checkbox: ")
                    finally:
                        driver.switch_to.default_content()
                except Exception as e:
                    self.log(f"[captcha] erro switch_to.frame: ")
            else:
                self.log("[captcha] iframe do recaptcha não encontrado")
                driver.get("https://site-com-recaptcha.com")

            
        
            # Se você usa Buster: a extensão pode interagir com o desafio automaticamente.
            # Aqui aguardamos o token aparecer no DOM (g-recaptcha-response ou input#recaptcha-token)
            self.log("[captcha] aguardando token resolver (sem timeout)...")
           
            token = None
            timeout = self.obter_timeout()  # segundos mudar para aumetar o tempo de espera para resolução do captcha
            inicio = time.time()

            while time.time() - inicio < timeout:
                try:
                    # 1) input hidden recaptcha-token (algumas implementações do INPI colocam o token aqui)
                   
                    try:
                        token_input = driver.find_element(By.ID, "recaptcha-token")
                        val = token_input.get_attribute("value")
                        if val and len(val) > 10:
                            token = val
                            break
                    except NoSuchElementException:
                        pass

                    # 2) textarea.g-recaptcha-response
                    try:
                        gr = driver.find_element(By.CSS_SELECTOR, "textarea.g-recaptcha-response")
                        val2 = gr.get_attribute("value")
                        if val2 and len(val2) > 10:
                            token = val2
                            break
                    except NoSuchElementException:
                        pass

                except Exception as e:
                    print("debug token check:")
                time.sleep(1)

            if token:
                self.log(f"[captcha] token detectado (len={len(token)})")
                # injeta token em possíveis campos e aciona o botão download
                try:
                    # tentar preencher g-recaptcha-response via JS (algumas páginas aceitam)
                    script_set = """
                    (function(t){
                        var ga = document.querySelector('textarea.g-recaptcha-response');
                        if(ga){ ga.value = t; ga.style.display='block'; }
                        var ri = document.getElementById('recaptcha-token');
                        if(ri) ri.value = t;
                        return true;
                    })(arguments[0]);
                    """
                    driver.execute_script(script_set, token)
                except Exception as e:
                    print("Erro ao injetar token via JS:")

                # clica no botão #captchaButton (Download)
                try:
                    btn = WebDriverWait(driver,1).until(
                        EC.element_to_be_clickable((By.ID, "captchaButton"))
                    )
                    btn.click()
                    self.log("[captcha] botão Download clicado")
                # aguarda o PDF ser baixado
                    selenium = self._selenium_por_worker(worker_id)
                    caminho_pdf = selenium.wait_for_download(worker_id)


                    if caminho_pdf:
                        self._renomear_pdf_para_processo(worker_id, caminho_pdf)

                    else:
                        self.log("❌ PDF não foi baixado.")
                        self._finalizar_processo_atual
    
                except Exception as e:
                    self.log(f"[captcha] falha ao clicar #captchaButton: {e}")
            else:
                self.log("[captcha] token não detectado — talvez Buster não resolveu automaticamente.")

           
                self.log("[captcha] nenhum PDF detectado no diretório de download (timeout).")
        except Exception as e:
            self.log(f"[captcha] erro no tratamento do modal: {e}")
                

    def _renomear_pdf_para_processo(self, worker_id, caminho_pdf):
        try:
            numero = self.numero_atual.get(worker_id)
            if not numero:
                self.log(
                    f"⚠️ Worker {worker_id} — Processo atual não definido para renomear PDF."
                )
                return
    
            # sanitiza o número (remove caracteres inválidos)
            nome_seguro = "".join(c for c in str(numero) if c.isalnum())
    
            # pasta correta do worker (garantido)
            pasta_worker = Path(self.download_dirs[worker_id])
    
            # garante que a pasta existe (proteção extra)
            pasta_worker.mkdir(parents=True, exist_ok=True)
    
            novo_nome = f"{nome_seguro}.pdf"
            novo_caminho = pasta_worker / novo_nome
    
            # se já existir, adiciona timestamp
            if novo_caminho.exists():
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                novo_caminho = pasta_worker / f"{nome_seguro}_{ts}.pdf"
    
            # caminho_pdf pode ser str ou Path
            caminho_pdf = Path(caminho_pdf)
    
            # renomeia (mesma pasta, sem mover entre workers)
            shutil.move(str(caminho_pdf), str(novo_caminho))
    
            self.log(
                f"📄 Worker {worker_id} — PDF renomeado para: {novo_caminho.name}"
            )

            # ✅ Verifica se o PDF realmente existe antes de registrar
            if novo_caminho.exists():
                self._registrar_processo_concluido(numero)
                QTimer.singleShot(
                    0, lambda: self._finalizar_processo_atual(worker_id)
                )
            else:
                self.log(
                    f"⚠️ Worker {worker_id} — PDF não encontrado após renomear, processo NÃO registrado."
                )

    
        except Exception as e:
            self.log(
                f"❌ Worker {worker_id} — Erro ao renomear PDF: {e}"
            )
  


  

    def selenium_get_recaptcha_iframe(self, driver):
        # procura iframe que contém 'anchor' (checkbox) ou 'api2/anchor'
        try:
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            for fr in iframes:
                src = fr.get_attribute("src") or ""
                if "api2/anchor" in src or "recaptcha" in src and "anchor" in src:
                    return fr
            # fallback: iframe title containing "reCAPTCHA"
            for fr in iframes:
                title = fr.get_attribute("title") or ""
                if "reCAPTCHA" in title or "recaptcha" in title.lower():
                    return fr
        except Exception:
            pass
        return None 
           
    def _pdf_baixado_com_sucesso(self, worker_id, caminho_final):
        caminho_pdf = os.path.abspath(caminho_final)

        self.log(f"📥 Worker {worker_id} — PDF salvo: {caminho_pdf}")

        QTimer.singleShot(
            0, lambda: self._finalizar_processo_atual(worker_id)
        )



    
    def _atualizar_contador_ui(self):
        texto = f"Processos extraídos: {self.processos_extraidos} / {self.total_processos}"
        self.label_contador.setText(texto)
        
    def _finalizar_processo_atual(self, worker_id):
        self.processando[worker_id] = False
        self.numero_atual[worker_id] = None
        self.processos_extraidos += 1
        self._atualizar_contador_ui()

        QTimer.singleShot(
            0, lambda: self._processar_proximo(worker_id)
        )



                                     
                        
    def closeEvent(self, event):
        # encerra selenium quando fechar UI
        try:
            self.selenium.stop()
        except Exception:
            pass
        event.accept()


# Se você quer testar este arquivo standalone (sem main.py),
# comente a importação em main.py e execute este módulo diretamente.
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    w = MainApp()
    w.show()
    sys.exit(app.exec_())
